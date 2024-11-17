# Generating an Excel report from GeoServer layers and data using Python
# Establishing a connection to GeoServer using specified credentials and retrieving relevant data
# Formatting the Excel report with specified styles for readability
# The report includes:
# - Detailed layer metadata and structure from the GeoServer catalog
# - Styling for cells with custom colors and fonts to highlight important data
# - Error handling for connection and data retrieval issues
# Cells are formatted with:
# - Dark blue font for 'N/A' values
# - Light blue background for cells containing descriptive metadata or status indicators


import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from geoserver.catalog import Catalog
import requests
import re

# GeoServer connection parameters
geoserver_url = "https://your-geoserver-url/geoserver/rest"  # Replace with your GeoServer URL
username = "your_username"  # Replace with your username
password = "your_password"  # Replace with your password

# Styling for cells
dark_blue_font = Font(color='00008B')
light_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Light blue fill

# Connect to the GeoServer catalog using gsconfig
cat = Catalog(geoserver_url, username, password)

def fetch_workspace_details():
    """Fetch detailed information about all workspaces."""
    workspaces = []
    for ws in cat.get_workspaces():
        workspaces.append({
            'name': ws.name,
            'href': getattr(ws, 'href', 'N/A')  # Safely getting href if exists
        })
    print("Workspace details fetched.")
    return workspaces

def fetch_store_details():
    """Fetch details for all data stores, including workspace name."""
    stores = []
    for store in cat.get_stores():
        store_url = f"{geoserver_url}/workspaces/{store.workspace.name}/datastores/{store.name}.html"
        stores.append({
            'workspace_name': store.workspace.name,
            'store_name': store.name,
            'store_url': store_url
        })
    stores = sorted(stores, key=lambda x: (x['workspace_name'], x['store_name']))  # Sorting by Workspace and Store
    print("Store details fetched and sorted.")
    return stores

def fetch_default_style(layer_name):
    """Fetch the default style for a given layer using the GeoServer REST API."""
    layer_url = f"{geoserver_url}/layers/{layer_name}.json"
    response = requests.get(layer_url, auth=(username, password))
    
    if response.status_code == 200:  # Fixed unmatched ')' issue
        layer_data = response.json()
        style_info = layer_data.get("layer", {}).get("defaultStyle", {}).get("name", "N/A")
        return style_info
    else:
        return "N/A"

def fetch_available_styles(layer_name):
    """Fetch available styles for a given layer from the GeoServer REST API."""
    layer_url = f"{geoserver_url}/layers/{layer_name}.json"
    response = requests.get(layer_url, auth=(username, password))
    
    if response.status_code == 200:
        layer_data = response.json()
        
        # Check if 'styles' is a valid dictionary and contains a 'style' list
        styles_data = layer_data.get("layer", {}).get("styles", {})
        
        if isinstance(styles_data, dict) and "style" in styles_data:
            available_styles = [style.get('name', 'N/A') for style in styles_data["style"] if isinstance(style, dict)]
            return ', '.join(available_styles) if available_styles else 'N/A'
        
    return "N/A"

def extract_epsg_code(bbox):
    """Extract EPSG code from the bounding box."""
    match = re.search(r'EPSG:\d+', str(bbox))
    if match:
        return match.group(0)
    return 'N/A'

def fetch_group_details():
    """Fetch detailed information about all layer groups."""
    groups = []
    group_to_layers = {}
    
    for group in cat.get_layergroups():
        workspace_name = getattr(group.workspace, 'name', 'N/A') if group.workspace else 'N/A'
        parent_name = getattr(group, 'name', 'N/A')
        
        # Get all possible layer group attributes
        bounds = getattr(group, 'bounds', 'N/A')
        title = getattr(group, 'title', 'N/A')
        mode = getattr(group, 'mode', 'N/A')
        layers = ', '.join([layer.name if hasattr(layer, 'name') else layer for layer in group.layers])
        
        # Extract the EPSG code from bounds
        crs = extract_epsg_code(str(bounds))
        
        # Collect group details
        groups.append({
            'workspace_name': workspace_name,
            'group_name': parent_name,
            'title': title,
            'crs': crs,
            'bounds': str(bounds),
            'mode': mode,
            'layers': layers
        })
        
        # Associate layers with the group for later use
        for layer in group.layers:
            if isinstance(layer, str):
                group_to_layers[layer] = parent_name
            else:
                group_to_layers[layer.name] = parent_name

    groups = sorted(groups, key=lambda x: (x['group_name']))  # Sorting by Group
    print("Layer group details fetched and sorted.")
    return groups, group_to_layers

def fetch_layer_details(group_to_layers):
    """Fetch detailed information about all layers, including their parent group."""
    layers = []
    
    for layer in cat.get_layers():
        resource = layer.resource
        parent_group = group_to_layers.get(layer.name, 'N/A')  # Get the parent group name if exists
        bbox = getattr(resource, 'latlon_bbox', 'N/A')  # Bounding box
        crs = extract_epsg_code(bbox)
        
        # Fetch default style and available styles for the layer
        default_style = fetch_default_style(layer.name)
        available_styles = fetch_available_styles(layer.name)
        
        layers.append({
            'workspace_name': resource.store.workspace.name if resource.store else 'N/A',
            'store': getattr(resource.store, 'name', 'N/A') if resource.store else 'N/A',
            'group_name': parent_group,
            'name': layer.name,
            'title': getattr(resource, 'title', 'N/A'),
            'default_style': default_style,
            'available_styles': available_styles,
            'crs': crs,
            'bbox': bbox,
            'abstract': getattr(resource, 'abstract', 'N/A')
        })
    layers = sorted(layers, key=lambda x: (x['workspace_name'], x['store'], x['group_name'], x['name']))  # Sorting by Workspace, Store, Group, and Layer Name
    print("Layer details fetched and sorted.")
    return layers

def fetch_styles():
    """Fetch all styles from the GeoServer REST API."""
    styles_url = f"{geoserver_url}/workspaces/cgs/styles.json"
    response = requests.get(styles_url, auth=(username, password))
    
    if response.status_code == 200:
        styles_data = response.json()
        styles_list = styles_data.get("styles", {}).get("style", [])
        return styles_list
    return []

def create_group_worksheets(wb, groups, layers):
    """Create worksheets for each layer group and fill them with layer and group details."""
    for group in groups:
        group_name = group['group_name']
        ws = wb.create_sheet(title=f"Group {group_name[:25]}")  # Truncate long names to 25 chars
        ws.sheet_properties.tabColor = "ADD8E6"  # Apply light blue tab color for layer groups
        
        # Add headers to the worksheet
        ws.append(["Workspace", "Store", "Group", "Group title", "Type", "Layer", "Child title", "Style"])
        
        # Populate rows for the group
        for layer_name in group['layers'].split(', '):
            matching_layer = next((layer for layer in layers if layer['name'] == layer_name), None)
            if matching_layer:
                # Populate row for each layer in the group
                ws.append([
                    matching_layer['workspace_name'],
                    matching_layer['store'],
                    group_name,
                    group['title'],  # Populate Group title
                    "Layer",
                    matching_layer['name'],
                    matching_layer['title'],  # Populate Child title with layer title
                    matching_layer['default_style']
                ])
            else:
                # If no matching layer is found, it's assumed to be another group
                ws.append([
                    group['workspace_name'],
                    "N/A",
                    group_name,
                    group['title'],  # Populate Group title
                    "Layer Group",
                    layer_name,
                    "N/A",  # Child title set to N/A if no layer title
                    "N/A"
                ])
        
        # Populate 'Child title' for Layer Groups by matching Layer and Group Name
        populate_child_title(wb, ws)

        # Add header hyperlinks to 'Workspace', 'Store', 'Group', 'Layer', and 'Style' columns
        add_header_hyperlinks(ws)

        # Link 'Workspace', 'Store', 'Group', 'Layer', and 'Style' columns with their respective sheets
        link_columns_with_sheets(ws, wb)

        # Create hyperlinks where 'Type' is 'Layer Group'
        add_layer_group_hyperlinks(wb, ws)

        # Apply the formatting functions to each group worksheet
        format_worksheet(ws)
        adjust_column_width(ws)
        apply_na_color(ws)
        print(f"Worksheet for Group {group_name} created.")

def populate_child_title(wb, ws):
    """Populate 'Child title' for rows where 'Type' = 'Layer Group' by matching 'Layer' and 'Group Name'."""
    layer_groups_ws = wb['Layer Groups']  # Get the Layer Groups worksheet
    layer_group_map = {row[0]: row[1] for row in layer_groups_ws.iter_rows(min_row=2, values_only=True)}  # Mapping 'Group Name' to 'Title'

    for row in ws.iter_rows(min_row=2):
        if row[4].value == "Layer Group":  # Check if 'Type' is 'Layer Group'
            layer_name = row[5].value.replace('cgs:', '')  # Remove 'cgs:' prefix
            if layer_name in layer_group_map:
                row[6].value = layer_group_map[layer_name]  # Populate 'Child title' with the matched 'Title'

def add_layer_group_hyperlinks(wb, ws):
    """Add hyperlinks where 'Type' = 'Layer Group' to 'Group Name' by matching the 'Layer' column."""
    layer_groups_ws = wb['Layer Groups']  # Get the 'Layer Groups' worksheet
    group_name_map = {row[0]: f"A{idx+2}" for idx, row in enumerate(layer_groups_ws.iter_rows(min_row=2, values_only=True))}  # Map 'Group Name' to row in Layer Groups
    
    for row in ws.iter_rows(min_row=2):
        if row[4].value == "Layer Group":  # Check if 'Type' is 'Layer Group'
            layer_name = row[5].value.replace('cgs:', '')  # Remove 'cgs:' prefix from Layer column
            if layer_name in group_name_map:
                row[5].hyperlink = f"#'Layer Groups'!{group_name_map[layer_name]}"  # Add hyperlink to the Group Name cell in 'Layer Groups'
                row[5].style = "Hyperlink"  # Set hyperlink style

def add_header_hyperlinks(ws):
    """Add header hyperlinks to the 'Workspace', 'Store', 'Group', 'Layer', and 'Style' columns."""
    ws.cell(row=1, column=1).hyperlink = "#'Workspaces'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    
    ws.cell(row=1, column=2).hyperlink = "#'Stores'!A1"
    ws.cell(row=1, column=2).style = "Hyperlink"
    
    ws.cell(row=1, column=3).hyperlink = "#'Layer Groups'!A1"
    ws.cell(row=1, column=3).style = "Hyperlink"

    # Add hyperlink for the 'Layer' header to link to the 'Layers' worksheet
    ws.cell(row=1, column=6).hyperlink = "#'Layers'!A1"
    ws.cell(row=1, column=6).style = "Hyperlink"

    # Add hyperlink for the 'Style' header to link to the 'Styles' worksheet
    ws.cell(row=1, column=8).hyperlink = "#'Styles'!A1"
    ws.cell(row=1, column=8).style = "Hyperlink"

def add_header_hyperlinks_layers(ws):
    """Add header hyperlinks to the 'Workspace', 'Store', 'Group', and 'Default style' columns in the 'Layers' worksheet."""
    ws.cell(row=1, column=1).hyperlink = "#'Workspaces'!A1"
    ws.cell(row=1, column=1).style = "Hyperlink"
    
    ws.cell(row=1, column=2).hyperlink = "#'Stores'!A1"
    ws.cell(row=1, column=2).style = "Hyperlink"
    
    ws.cell(row=1, column=3).hyperlink = "#'Layer Groups'!A1"
    ws.cell(row=1, column=3).style = "Hyperlink"
    
    # Add hyperlink for the 'Default style' header to link to the 'Styles' worksheet
    ws.cell(row=1, column=6).hyperlink = "#'Styles'!A1"
    ws.cell(row=1, column=6).style = "Hyperlink"

def add_header_hyperlinks_layer_groups(ws):
    """Add hyperlink to the 'Layers' header in the 'Layer Groups' worksheet."""
    ws.cell(row=1, column=6).hyperlink = "#'Layers'!A1"
    ws.cell(row=1, column=6).style = "Hyperlink"

def link_columns_with_sheets(ws, wb):
    """Add hyperlinks to the 'Workspace', 'Store', 'Group', 'Layer', and 'Style' columns in group worksheets."""
    # Create lookups
    workspace_lookup = {row[0]: f"Workspaces!A{idx+2}" for idx, row in enumerate(wb["Workspaces"].iter_rows(min_row=2, values_only=True))}
    store_lookup = {row[1]: f"Stores!B{idx+2}" for idx, row in enumerate(wb["Stores"].iter_rows(min_row=2, values_only=True))}
    group_lookup = {row[0]: f"A{idx+2}" for idx, row in enumerate(wb["Layer Groups"].iter_rows(min_row=2, values_only=True))}
    layer_lookup = {row[3]: f"Layers!D{idx+2}" for idx, row in enumerate(wb["Layers"].iter_rows(min_row=2, values_only=True))}
    style_lookup = {row[0]: f"Styles!A{idx+2}" for idx, row in enumerate(wb["Styles"].iter_rows(min_row=2, values_only=True))}

    # Apply hyperlinks for each row in the current group worksheet
    for row in ws.iter_rows(min_row=2):
        workspace_name = row[0].value
        store_name = row[1].value
        group_name = row[2].value
        layer_name = row[5].value
        style_name = row[7].value
        
        # Add hyperlink to 'Workspace'
        if workspace_name in workspace_lookup:
            row[0].hyperlink = f"#{workspace_lookup[workspace_name]}"
            row[0].style = "Hyperlink"
        
        # Add hyperlink to 'Store'
        if store_name in store_lookup:
            row[1].hyperlink = f"#{store_lookup[store_name]}"
            row[1].style = "Hyperlink"
        
        # Add hyperlink to 'Group'
        if group_name in group_lookup:
            row[2].hyperlink = f"#'Layer Groups'!{group_lookup[group_name]}"
            row[2].style = "Hyperlink"
        
        # Add hyperlink to 'Layer'
        if layer_name in layer_lookup:
            row[5].hyperlink = f"#{layer_lookup[layer_name]}"
            row[5].style = "Hyperlink"

        # Add hyperlink to 'Style'
        if style_name in style_lookup:
            row[7].hyperlink = f"#{style_lookup[style_name]}"
            row[7].style = "Hyperlink"

def adjust_column_width(sheet):
    """Adjust the column widths based on the maximum length of the data in each column, but limit to 50."""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column letter

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = min(max_length + 2, 50)  # Limit to 50
        sheet.column_dimensions[column].width = adjusted_width

def format_worksheet(sheet):
    """Freeze the first row, add filters to all columns, and set zoom to 125%."""
    sheet.freeze_panes = "A2"  # Freeze the first row
    sheet.auto_filter.ref = sheet.dimensions  # Set the auto-filter
    sheet.sheet_view.zoomScale = 125  # Set zoom to 125%

def apply_na_color(sheet):
    """Apply dark blue font color to all 'N/A' values in the worksheet."""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == 'N/A':
                cell.font = dark_blue_font

def write_to_excel(workspaces, stores, groups, layers, styles, output_file):
    wb = Workbook()

    # Workspaces sheet
    ws_workspaces = wb.active
    ws_workspaces.title = "Workspaces"
    ws_workspaces.append(["Workspace Name", "HREF"])
    for workspace in workspaces:
        ws_workspaces.append([workspace['name'], workspace['href']])
    format_worksheet(ws_workspaces)
    adjust_column_width(ws_workspaces)
    apply_na_color(ws_workspaces)
    print("Workspaces worksheet completed.")

    # Stores sheet (added second after 'Workspaces')
    ws_stores = wb.create_sheet(title="Stores", index=1)
    ws_stores.append(["Workspace Name", "Store Name", "Store URL"])
    
    # Fetch the link locations from the Workspaces sheet
    workspace_lookup = {workspace['name']: f"Workspaces!A{idx+2}" for idx, workspace in enumerate(workspaces)}
    
    for store in stores:
        ws_stores.append([store['workspace_name'], store['store_name'], store['store_url']])
        # Create a hyperlink to the Workspaces worksheet for the Workspace Name
        if store['workspace_name'] in workspace_lookup:
            link_cell = ws_stores.cell(row=ws_stores.max_row, column=1)
            link_cell.hyperlink = f"#{workspace_lookup[store['workspace_name']]}"
            link_cell.style = "Hyperlink"
    
    format_worksheet(ws_stores)
    adjust_column_width(ws_stores)
    apply_na_color(ws_stores)
    print("Stores worksheet completed.")

    # Groups sheet with 'CRS' and other details included
    ws_groups = wb.create_sheet(title="Layer Groups", index=2)
    ws_groups.append(["Group Name", "Title", "CRS", "Bounds", "Mode", "Layers"])
    
    group_worksheets = {}
    
    for group in groups:
        ws_groups.append([group['group_name'], group['title'], group['crs'], group['bounds'], group['mode'], group['layers']])
        
        # Store the reference for hyperlinks
        group_worksheets[group['group_name']] = f"Group {group['group_name'][:25]}"

    format_worksheet(ws_groups)
    adjust_column_width(ws_groups)
    apply_na_color(ws_groups)
    
    # Add hyperlinks to 'Group Name' column in 'Layer Groups'
    for idx, group_name in enumerate(group_worksheets.keys(), start=2):
        cell = ws_groups.cell(row=idx, column=1)
        # Use group name within single quotes and parentheses
        cell.hyperlink = f"#'Group {group_name[:25]}'!A1"
        cell.style = "Hyperlink"

    # Add hyperlink to the 'Layers' header in 'Layer Groups'
    add_header_hyperlinks_layer_groups(ws_groups)

    print("Layer Groups worksheet completed.")

    # Layers sheet (added after 'Layer Groups')
    ws_layers = wb.create_sheet(title="Layers", index=3)
    ws_layers.append(["Workspace Name", "Store Name", "Group Name", "Layer Name", "Child title", "Default style", "Available styles", "CRS", "Bounding Box", "Abstract"])
    
    for layer in layers:
        ws_layers.append([layer['workspace_name'], layer['store'], layer['group_name'], layer['name'], layer['title'], layer['default_style'], layer['available_styles'], layer['crs'], str(layer['bbox']), layer['abstract']])

    # Add header hyperlinks to 'Layers' sheet
    add_header_hyperlinks_layers(ws_layers)

    format_worksheet(ws_layers)
    adjust_column_width(ws_layers)
    apply_na_color(ws_layers)
    print("Layers worksheet completed.")

    # Styles sheet (added right after 'Layers')
    ws_styles = wb.create_sheet(title="Styles", index=4)
    ws_styles.append(["Style name", "Style link"])
    
    for style in styles:
        style_name = style.get("name", "N/A")
        style_link = f"https://csg-geoserver.university.innopolis.ru/geoserver/rest/workspaces/cgs/styles/{style_name}.html"
        sld_link = f"https://csg-geoserver.university.innopolis.ru/geoserver/rest/workspaces/cgs/styles/{style_name}.sld"
        
        style_name_cell = ws_styles.cell(row=ws_styles.max_row + 1, column=1)
        style_name_cell.value = style_name
        style_name_cell.hyperlink = sld_link
        style_name_cell.style = "Hyperlink"
        
        style_link_cell = ws_styles.cell(row=ws_styles.max_row, column=2)
        style_link_cell.value = style_link
        style_link_cell.hyperlink = style_link  # Add hyperlink using the cell value
    
    format_worksheet(ws_styles)
    adjust_column_width(ws_styles)
    apply_na_color(ws_styles)
    print("Styles worksheet completed.")

    # Create worksheets for each layer group
    create_group_worksheets(wb, groups, layers)

    # Set 'Layer Groups' as the default active sheet
    wb.active = wb["Layer Groups"]

    # Save the workbook
    wb.save(output_file)
    print(f"Data has been written to {output_file}")

def generate_filename():
    """Generate a filename based on the current timestamp."""
    timestamp = datetime.now().strftime("%Y%m%d_%H-%M-%S")
    return f"geoserver_data_{timestamp}.xlsx"

if __name__ == "__main__":
    # Fetch detailed information
    workspaces = fetch_workspace_details()
    stores = fetch_store_details()  # Fetch store details
    groups, group_to_layers = fetch_group_details()  # Fetch group details and map layers to groups
    layers = fetch_layer_details(group_to_layers)  # Fetch layer details and associate them with groups
    styles = fetch_styles()  # Fetch styles for the Styles worksheet

    # Generate dynamic filename
    output_filename = generate_filename()

    # Write data to Excel file
    write_to_excel(workspaces, stores, groups, layers, styles, output_filename)
